"""Serviço do Resumo Operacional — KPIs, gráficos e tabela analítica."""

from __future__ import annotations

from datetime import datetime
from typing import Any

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from core.cache_read import TTL_DASHBOARD
from core.styles import COLORS
from core.theme import (
    PLOTLY_AXIS_SIZE,
    PLOTLY_BAR_LABEL_SIZE,
    PLOTLY_FONT,
    PLOTLY_TITLE_SIZE,
)
from core.tratativa_constants import TRATATIVA_FILTRO_TODOS
from repositories.resumo_operacional_repository import (
    FILTRO_TODOS_CLIENTE,
    FILTRO_TODOS_MOTIVO,
    FiltrosResumo,
    listar_agregado_tabela,
    listar_opcoes_cliente,
    listar_opcoes_motivo,
    obter_kpis,
    top5_clientes_quantidade,
    top5_clientes_valor,
    top5_motivos_quantidade,
    top5_motivos_valor,
)
from services.dashboard_service import (
    MESES_LABEL,
    CHART_GRID,
    CHART_HEIGHT,
    COR_DEVOLUCOES,
    COR_FINANCEIRO,
    _layout_plotly,
    _rotulo_int,
    _rotulo_moeda,
    mes_label_para_numero,
    obter_anos_disponiveis,
    obter_periodo_padrao,
)
from services.devolucao_service import _formatar_valor_br

COR_RESUMO_QTD = COR_DEVOLUCOES
COR_RESUMO_VALOR = COR_FINANCEIRO

MODO_MOTIVOS = "Motivos"
MODO_CLIENTES = "Clientes"
MODOS_TABELA = (MODO_MOTIVOS, MODO_CLIENTES)

ORDENAR_QUANTIDADE = "Quantidade"
ORDENAR_VALOR = "Valor"
MODOS_ORDENACAO = (ORDENAR_QUANTIDADE, ORDENAR_VALOR)


def _classificar_por_valor(ordenacao: str) -> bool:
    return (ordenacao or ORDENAR_QUANTIDADE).strip() == ORDENAR_VALOR


def montar_filtros(
    mes: int,
    ano: int,
    cliente: str,
    motivo: str,
    tratativa: str,
) -> FiltrosResumo:
    return FiltrosResumo(
        mes=int(mes),
        ano=int(ano),
        cliente=(cliente or FILTRO_TODOS_CLIENTE).strip(),
        motivo=(motivo or FILTRO_TODOS_MOTIVO).strip(),
        tratativa=(tratativa or TRATATIVA_FILTRO_TODOS).strip(),
    )


def _formatar_media_movel_diaria(valor: float) -> str:
    return f"{_formatar_valor_br(valor)} / dia"


def _cards_resumo_formatados(raw: dict[str, Any]) -> dict[str, str]:
    return {
        "total_devolucoes": str(int(raw.get("total_devolucoes") or 0)),
        "valor_total": _formatar_valor_br(raw.get("soma_valor_nf", 0)),
        "ticket_medio": _formatar_valor_br(raw.get("ticket_medio", 0)),
        "media_movel": _formatar_media_movel_diaria(float(raw.get("media_movel_diaria") or 0)),
    }


def _barra_horizontal(
    categorias: list[str],
    valores: list[float],
    textos: list[str],
    cor: str,
    hovertemplate: str,
) -> go.Bar:
    cats = list(reversed(categorias))
    vals = list(reversed(valores))
    txt = list(reversed(textos))
    return go.Bar(
        orientation="h",
        y=cats,
        x=vals,
        text=txt,
        textposition="outside",
        textfont=dict(color="#e6edf3", size=PLOTLY_BAR_LABEL_SIZE, family=PLOTLY_FONT),
        marker=dict(color=cor, line=dict(width=0), cornerradius=8),
        width=0.52,
        hovertemplate=hovertemplate,
        cliponaxis=False,
    )


def _layout_horizontal(titulo: str) -> dict[str, Any]:
    layout = _layout_plotly(titulo, height=CHART_HEIGHT)
    layout["margin"] = dict(l=8, r=48, t=52, b=18)
    layout["yaxis"] = dict(
        visible=True,
        showticklabels=True,
        showgrid=False,
        gridcolor=CHART_GRID,
        zeroline=False,
        showline=False,
        tickfont=dict(color=COLORS["text_muted"], size=PLOTLY_AXIS_SIZE, family=PLOTLY_FONT),
        automargin=True,
    )
    layout["xaxis"] = dict(
        visible=False,
        showticklabels=False,
        showgrid=True,
        gridcolor=CHART_GRID,
        zeroline=False,
        showline=False,
    )
    return layout


def _finalizar_figura_horizontal(fig: go.Figure, titulo: str) -> go.Figure:
    fig.update_layout(**_layout_horizontal(titulo))
    fig.update_xaxes(
        visible=False,
        showticklabels=False,
        showgrid=True,
        gridcolor=CHART_GRID,
        zeroline=False,
    )
    fig.update_yaxes(
        visible=True,
        showticklabels=True,
        showgrid=False,
        automargin=True,
        tickfont=dict(color=COLORS["text_muted"], size=PLOTLY_AXIS_SIZE, family=PLOTLY_FONT),
    )
    return fig


def _grafico_top5_motivos_qtd(filtros: FiltrosResumo) -> go.Figure:
    rows = top5_motivos_quantidade(filtros)
    labels = [r[0] for r in rows]
    valores = [float(r[1]) for r in rows]
    textos = [_rotulo_int(v) for v in valores]
    fig = go.Figure(
        data=[
            _barra_horizontal(
                labels,
                valores,
                textos,
                COR_RESUMO_QTD,
                "%{y}<br><b>%{x}</b> devoluções<extra></extra>",
            )
        ]
    )
    return _finalizar_figura_horizontal(fig, "Top 5 Motivos — Quantidade")


def _grafico_top5_motivos_valor(filtros: FiltrosResumo) -> go.Figure:
    rows = top5_motivos_valor(filtros)
    labels = [r[0] for r in rows]
    valores = [float(r[1]) for r in rows]
    textos = [_rotulo_moeda(v, decimais=False) for v in valores]
    fig = go.Figure(
        data=[
            _barra_horizontal(
                labels,
                valores,
                textos,
                COR_RESUMO_VALOR,
                "%{y}<br><b>R$ %{x:,.0f}</b><extra></extra>",
            )
        ]
    )
    return _finalizar_figura_horizontal(fig, "Top 5 Motivos — Valor")


def _grafico_top5_clientes_qtd(filtros: FiltrosResumo) -> go.Figure:
    rows = top5_clientes_quantidade(filtros)
    labels = [r["rotulo"] for r in rows]
    valores = [float(r["quantidade"]) for r in rows]
    textos = [_rotulo_int(v) for v in valores]
    fig = go.Figure(
        data=[
            _barra_horizontal(
                labels,
                valores,
                textos,
                COR_RESUMO_QTD,
                "%{y}<br><b>%{x}</b> devoluções<extra></extra>",
            )
        ]
    )
    return _finalizar_figura_horizontal(fig, "Top 5 Clientes — Quantidade")


def _grafico_top5_clientes_valor(filtros: FiltrosResumo) -> go.Figure:
    rows = top5_clientes_valor(filtros)
    labels = [r["rotulo"] for r in rows]
    valores = [float(r["valor"]) for r in rows]
    textos = [_rotulo_moeda(v, decimais=False) for v in valores]
    fig = go.Figure(
        data=[
            _barra_horizontal(
                labels,
                valores,
                textos,
                COR_RESUMO_VALOR,
                "%{y}<br><b>R$ %{x:,.0f}</b><extra></extra>",
            )
        ]
    )
    return _finalizar_figura_horizontal(fig, "Top 5 Clientes — Valor")


def preparar_tabela_motivos(rows: list[dict[str, Any]]) -> pd.DataFrame:
    colunas = ["Motivo", "Quantidade", "Valor Total", "Percentual"]
    if not rows:
        return pd.DataFrame(columns=colunas)
    linhas = [
        {
            "Motivo": r["motivo"],
            "Quantidade": int(r["quantidade"]),
            "Valor Total": _formatar_valor_br(r["valor_total"]),
            "Percentual": float(r["percentual"]),
        }
        for r in rows
    ]
    return pd.DataFrame(linhas, columns=colunas)


def preparar_tabela_clientes(rows: list[dict[str, Any]]) -> pd.DataFrame:
    colunas = ["Código", "Cliente", "Quantidade", "Valor Total", "Percentual"]
    if not rows:
        return pd.DataFrame(columns=colunas)
    linhas = [
        {
            "Código": r["cod_cliente"] or "—",
            "Cliente": r["cliente"] or "—",
            "Quantidade": int(r["quantidade"]),
            "Valor Total": _formatar_valor_br(r["valor_total"]),
            "Percentual": float(r["percentual"]),
        }
        for r in rows
    ]
    return pd.DataFrame(linhas, columns=colunas)


def nome_arquivo_resumo_operacional() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Resumo_Operacional_{ts}.xlsx"


def _parse_valor_total_excel(valor: Any) -> float | None:
    """Converte texto exibido na tela (R$ 1.234,56) para número do Excel."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto or texto == "—":
        return None
    limpo = (
        texto.replace("R$", "")
        .replace("\xa0", "")
        .strip()
        .replace(".", "")
        .replace(",", ".")
    )
    try:
        return float(limpo)
    except ValueError:
        return None


def _preparar_df_exportacao_tabela_resumo(df: pd.DataFrame, modo: str) -> pd.DataFrame:
    """Prepara o DataFrame exibido para exportação Excel (sem heat map)."""
    if modo == MODO_CLIENTES:
        colunas = ["Código Cliente", "Cliente", "Quantidade", "Valor Total", "Percentual"]
    else:
        colunas = ["Motivo", "Quantidade", "Valor Total", "Percentual"]

    if df.empty:
        return pd.DataFrame(columns=colunas)

    export = df.copy()
    if modo == MODO_CLIENTES and "Código" in export.columns:
        export = export.rename(columns={"Código": "Código Cliente"})

    export["Quantidade"] = pd.to_numeric(export["Quantidade"], errors="coerce")
    export["Valor Total"] = export["Valor Total"].apply(_parse_valor_total_excel)
    export["Percentual"] = pd.to_numeric(export["Percentual"], errors="coerce") / 100.0

    return export[colunas]


def export_tabela_resumo_excel_bytes(df: pd.DataFrame, modo: str) -> bytes:
    """Gera XLSX da tabela analítica exibida na tela."""
    from io import BytesIO

    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    export_df = _preparar_df_exportacao_tabela_resumo(df, modo)
    buffer = BytesIO()
    sheet_name = "Resumo"

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        if export_df.shape[0] > 0 or export_df.shape[1] > 0:
            ws.auto_filter.ref = ws.dimensions

        colunas = {name: idx + 1 for idx, name in enumerate(export_df.columns)}
        ultima_linha = len(export_df) + 1

        if "Quantidade" in colunas:
            col = colunas["Quantidade"]
            for row in range(2, ultima_linha + 1):
                ws.cell(row=row, column=col).number_format = "0"

        if "Valor Total" in colunas:
            col = colunas["Valor Total"]
            for row in range(2, ultima_linha + 1):
                ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'

        if "Percentual" in colunas:
            col = colunas["Percentual"]
            for row in range(2, ultima_linha + 1):
                ws.cell(row=row, column=col).number_format = "0.0%"

        for idx, col_name in enumerate(export_df.columns, start=1):
            letter = get_column_letter(idx)
            max_len = len(str(col_name))
            for row in range(2, ultima_linha + 1):
                valor = ws.cell(row=row, column=idx).value
                if valor is not None:
                    max_len = max(max_len, len(str(valor)))
            ws.column_dimensions[letter].width = min(max_len + 2, 48)

    buffer.seek(0)
    return buffer.getvalue()


@st.cache_data(ttl=TTL_DASHBOARD, show_spinner=False)
def obter_opcoes_filtros_cache(mes: int, ano: int) -> dict[str, Any]:
    clientes = listar_opcoes_cliente(mes, ano)
    motivos = listar_opcoes_motivo(mes, ano)
    return {
        "clientes": tuple(clientes),
        "motivos": tuple(motivos),
    }


@st.cache_data(ttl=TTL_DASHBOARD, show_spinner=False)
def carregar_resumo_cache(
    mes: int,
    ano: int,
    cliente: str,
    motivo: str,
    tratativa: str,
) -> dict[str, Any]:
    filtros = montar_filtros(mes, ano, cliente, motivo, tratativa)
    kpis_raw = obter_kpis(filtros)
    return {
        "cards": _cards_resumo_formatados(kpis_raw),
        "filtros": filtros,
    }


@st.cache_data(ttl=TTL_DASHBOARD, show_spinner=False)
def obter_graficos_resumo_cache(
    mes: int,
    ano: int,
    cliente: str,
    motivo: str,
    tratativa: str,
) -> dict[str, dict[str, Any]]:
    filtros = montar_filtros(mes, ano, cliente, motivo, tratativa)
    return {
        "motivos_qtd": _grafico_top5_motivos_qtd(filtros).to_dict(),
        "motivos_valor": _grafico_top5_motivos_valor(filtros).to_dict(),
        "clientes_qtd": _grafico_top5_clientes_qtd(filtros).to_dict(),
        "clientes_valor": _grafico_top5_clientes_valor(filtros).to_dict(),
    }


def obter_graficos_resumo(
    mes: int,
    ano: int,
    cliente: str,
    motivo: str,
    tratativa: str,
) -> dict[str, go.Figure]:
    graficos = obter_graficos_resumo_cache(mes, ano, cliente, motivo, tratativa)
    return {chave: go.Figure(fig) for chave, fig in graficos.items()}


@st.cache_data(ttl=TTL_DASHBOARD, show_spinner=False)
def carregar_tabela_resumo_cache(
    mes: int,
    ano: int,
    cliente: str,
    motivo: str,
    tratativa: str,
    modo: str,
    ordenacao: str,
) -> pd.DataFrame:
    filtros = montar_filtros(mes, ano, cliente, motivo, tratativa)
    por_cliente = modo == MODO_CLIENTES
    por_valor = _classificar_por_valor(ordenacao)
    rows = listar_agregado_tabela(
        filtros,
        por_cliente=por_cliente,
        por_valor=por_valor,
    )
    if por_cliente:
        return preparar_tabela_clientes(rows)
    return preparar_tabela_motivos(rows)


__all__ = [
    "MESES_LABEL",
    "MODO_CLIENTES",
    "MODO_MOTIVOS",
    "MODOS_ORDENACAO",
    "MODOS_TABELA",
    "ORDENAR_QUANTIDADE",
    "ORDENAR_VALOR",
    "FILTRO_TODOS_CLIENTE",
    "FILTRO_TODOS_MOTIVO",
    "carregar_resumo_cache",
    "carregar_tabela_resumo_cache",
    "export_tabela_resumo_excel_bytes",
    "mes_label_para_numero",
    "montar_filtros",
    "nome_arquivo_resumo_operacional",
    "obter_anos_disponiveis",
    "obter_graficos_resumo",
    "obter_graficos_resumo_cache",
    "obter_opcoes_filtros_cache",
    "obter_periodo_padrao",
]
